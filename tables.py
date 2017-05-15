import wx
import wx.grid
import os
import time
from openpyxl import Workbook
from openpyxl import load_workbook
"""
This module is about extracting information from Excel to display in
a wxGrid using openpyxl. Can only read xlsx files.
Formulas are read as text and then saved as formulas again.
"""

class TABLES(object):
    """
    Reads data from Excel.
    """
    def __init__(self, other_self):
        self.other_self = other_self
        self.source = None #later this will be whatever source book we read.
        self.names = [] #list of loaded sheet names, for use in saving.
        
    def excel_to_grid(self, source, sheet, grid):
        """
        Opens the Excel file in source and loads the sheet into the grid.
        """
        self.source = source #save original workbook for later use when saving data.
        wb = load_workbook(source,data_only = False)

        names = wb.get_sheet_names()
        
        if sheet in names:
            self.names.append(sheet)
            grid.ClearGrid()
            if grid.GetNumberRows()>0:
                grid.DeleteRows(0,grid.GetNumberRows() ,True)
            if grid.GetNumberCols()>0:
                grid.DeleteCols(0,grid.GetNumberCols() ,True)

            
            sh = wb.get_sheet_by_name(name=sheet)
            num_rows = sh.max_row
            num_cols = sh.max_column
            self.SetGridRows(grid, num_rows)
            self.SetGridCols(grid, num_cols)#extra columns for results
            #print 'number of rows = ', num_rows
            #print 'number of columns = ', num_cols
            
            for curr_row in range(num_rows):
                for i in range(num_cols):
                    grid.SetCellValue(curr_row, i, str(sh.cell(row=curr_row+1,column=i+1).value))
            return True
        else:
            return False
        
    def grid_to_excel(self,target,grids):
        """Reads the grid and writes to the excel file, then saves.
        each input grid is a tuple, grid followed by sheet name"""
        #reading self.source, the original file.
        wb = load_workbook(self.source,data_only = False)
        sheets = [None]*len(grids) #array for all the sheets
        if len(grids)>0:
            for grid,num in zip(grids,range(len(grids)) ):
                #each "grid" is a tuple, element 0 is the grid, element 1 is the name
                sheets[num] = wb.get_sheet_by_name(name=grid[1])
                
                for r in range(grid[0].GetNumberRows()):
                    for c in range(grid[0].GetNumberCols()):
                        cell_value = grid[0].GetCellValue(r,c)
                        sh = sheets[num]
                        #save the value to the cell object's value attribute
                        sh.cell(row=r+1,column=c+1,value=cell_value)
            #save the workbook and we are done.
            wb.save(target)				
                    
    def SetGridRows(self, grid_name, no_of_rows):
        """
        Set grid, *grid_name*, to have rows, *no_of_rows*.
        """
        grid_name.ClearGrid() #clear all data first
        change_rows = no_of_rows - grid_name.GetNumberRows()
        if change_rows > 0:
                grid_name.AppendRows(change_rows) #always to end
        elif change_rows < 0:
                grid_name.DeleteRows(0, -change_rows) #from posn 0
        self.other_self.m_scrolledWindow3.SendSizeEvent() # make sure new size is fitted
                    
    def SetGridCols(self, grid_name, no_of_cols):
        """
        Set grid, *grid_name*, to have columns, *no_of_cols*.
        """
        grid_name.ClearGrid() #clear all data first
        change_cols = no_of_cols - grid_name.GetNumberCols()
        if change_cols > 0:
                grid_name.AppendCols(change_cols) #always to end
        elif change_cols < 0:
                grid_name.DeleteRows(0, -change_cols) #from posn 0
        self.other_self.m_scrolledWindow3.SendSizeEvent() # make sure new size is fitted



if __name__ == '__main__':
    app = wx.PySimpleApp()
    frame = wx.Frame(None, -1, size=(700,500), title = "Testing tables.py")
    grid = wx.grid.Grid(frame)
    grid.CreateGrid(20,6)
    b = '' #substitute for a namespace
    a = TABLES(b)
    a.excel_to_grid('RefStepE052.xlsx', 'Sheet1',grid)
    
    frame.Show(True)
    app.MainLoop()
