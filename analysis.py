"""
Analysis tool. Needs a file name to read (from the same folder) and a file name to save to (can be the same, then it updates) as well as a sheet name in which the data is stored.
Files must be of the format created by the main program graphframe.py>GridMaker.py.
Results are saved onto a new sheet, and dumped into the saving file. 
"""
from openpyxl import Workbook
from openpyxl import load_workbook
import warnings
import GTC

class Analyser(object):
    """
    Analysis object for a refstep table with arbitrary lengths of runs
    (not restricted to 10, but needs more than 1).
    """
    def __init__(self,book_name,sheet_name): 
        """
        Initialisation requires the name (change to path?) of the file, name of sheet to read
        and the 'window' to read data from. The window is a set of coordinates (y,y,x,x) of the
        top left and bottom right that define the block of data to load.
        """
        self.mean_col = 13 #column for the mean values
        self.std_col = self.mean_col+1
        self.DoF_col = self.mean_col-6 #might change if we add col for the actual range settings of the instruments for verification
        self.S_setting_col = self.DoF_col-3
        self.X_setting_col = self.S_setting_col-2 #should always be 2 cols away

        #warnings.simplefilter('ignore')
        self.wb = load_workbook(book_name,data_only = True)
        #warnings.simplefilter('default')
        #try command? if fail, let user know.
        self.sh = self.wb[sheet_name]
        
        self.sh_results = self.wb.create_sheet(title="Results")
        window = [1,self.sh.max_row+1,1,self.sh.max_column+1]
        self.data = self.get_data(window) #row,row,col,col

    def Save(self, name):
        """
        Saves the product excel sheet to a given name.
        """
        self.wb.save(name)
        
    def print_cell(self, value, row, column, sheet):
        """
        Prints a single cell to the sheet. function takes in the cell value,
        the cell row, and the cell column in this order
        """
        sheet.cell(row = row,column = column,value = value)
        #print(value,row,column)
        
    def PrintCol(self,col,start_row,start_col,sheet):
        """
        Prints an entire column using the print cell function. Requires the column to print, start row, start col.
        """
        for value, row in zip(col,range(start_row,len(col)+start_row)):
            self.print_cell(value,row,start_col,sheet)

    def read_cell(self,cell_row,cell_col):
        """
        Reads a cell from the loaded data. cell coordinates must be within the window specified earlier.
        function takes inputs as the cell row and cell column.
        """
        try:
            info = self.data[cell_row][cell_col]
        except:
            info =  0
            print("cannot read cell "+str((cell_row,cell_col)))
        return info

    def get_data(self,block_range):
        """
        Returns a 2D array of the entire window specified. the function requires the block range or window to be specified.
        """
        rows = []
        for i in range(block_range[0],block_range[1]):
            single_row = []
            for j in range(block_range[2],block_range[3]):
                single_row.append(self.sh.cell(row=i,column = j).value)
            rows.append(single_row)
        return rows

    def x_ratio(self,ureals,s0,s1):
        """
        Computes the ratios for a set of ascending GTC.ureal objects, for source x. The first two data points must belong to
        the range below and are used for the gain ratio computation.
        """
        #compute ratios for a set of ascending ureals, for source x. first two data points are for the gain ratio, ie they are on a lower range.
        print("lenght x "+str(len(ureals)))
        Sum = 0
        num = 0
        print([x.x for x in ureals])
        for j in range(2,(len(ureals)+1)/2):
            Sum = Sum +ureals[2*j-1]-ureals[2*j-2]
            num = num+1
        length = (len(ureals)-1)/2 #generalised instead of dividing later by 10.
        #print("sum "+str(Sum.x))
        #print("x in put was:")
        #print([x.x for x in ureals])
        lin_ratio = (1.0+1.0/(num*(s1-s0))*Sum) / (1.0+(1.0/(s1-s0))*(ureals[3]-ureals[2])) #computes linearity ratio? only if first and second readings are not the ones from previous range.
        lin_ratio.label = 'x linearity'
        gain_ratio = (1+1/(num*(s1-s0))*Sum)/(1+(1/(s1-s0))*(ureals[1]-ureals[0])) #computes gain ratio? only if first and second readings are not the ones from previous range.
        gain_ratio.label = 'x gain'
        return [lin_ratio,gain_ratio]
    
    def m_ratio(self,ureals,s0,s1):
        """
        Computes the ratios for a set of ascending GTC.ureal objects, for source x. The first two data points must belong to
        the range below and are used for the gain ratio computation.
        """
        #compute ratios for a set of ascending ureals, for the meter, first two data points are for the gain ratio, ie they are on a lower range.
        print("lenght m "+str(len(ureals)))
        Sum = 0
        num = 0
        for j in range(2,(len(ureals)+2)/2):
            Sum = Sum +ureals[2*j-1]-ureals[2*j-2]
            num = num+1
            #print("diff "+str(ureals[2*j-1].x-ureals[2*j-2].x))
        #print(num)
        #print("sum "+str(Sum.x))
        #print("m input was:")
        #print([x.x for x in ureals])
        lin_ratio = ((1.0/num)*Sum / (ureals[3]-ureals[2])) #computes linearity ratio, not using urealls[3]-ureals[2]. abs ensures either negative or positive is the same
        #previous line updated back to ureals[3]-ureals[2] since no longer sending the useless data point. why do we have it?
        lin_ratio.label = 'm linearity'
        gain_ratio = ((1.0/num)*Sum / (ureals[1]-ureals[0])) #computes gain ratio, if the gain ratio formula is the same as linearity just divided by the previous range elements. 
        gain_ratio.label = 'm gain'
        return [lin_ratio,gain_ratio]

    def split_set(self,ureals,s_settings):
        """
        Given a symmetric set of ascending and decending GTC.ureal objects that contain data for both meter and source
        this function splits the data up to its respective instruments (most points are used in both meter and source calculations)
        and then splits it further into ascending and decending components.
        """
        #split set of ascending and decending data to two ratio computations, for x and s
        length = len(ureals)
        top_data = ureals[:(length+1/2)] #include center point
        bottom_data = ureals[(length/2):]
        #meter and source require different subsets of each of these sets of data
        m_data = ureals[:2]+ureals[3:-3]+ureals[-2:] #first and second points for gain ratios, then from sixth point
        #meter calculations dont seem to ever make use of point ureals[2] so removed it, then
        #analysis is the same for even when doing a range that does not include a zero.
        #need to determine if the set includes that residual zero? or do we just remove it from all data sets...
        m_data = ureals[:2]+ureals[4:-4]+ureals[-2:]
        x_data = ureals[1:-1] #uses all data except first/last readings that are negative. 
        #these sequences need to be split to top and bottom sections, both of which include middle point.
        m_top = m_data[:(len(m_data)+1)/2]
        m_bottom = m_data[(len(m_data))/2:]
        x_top = x_data[:(len(x_data)+1)/2]
        x_bottom = x_data[len(x_data)/2:]

        s_bits = [s_settings[1],s_settings[2]] #s0 and s1 for this set. should be the same through the entire set?
        
        return [m_top,m_bottom,x_top,x_bottom]+s_bits

    def find_center(self,start_row):
        """
        Given a start row, reads through the test column (column 2, settings column for the source)
        and identifies the mid points (3 repeated settings). returns the mid point row and end point row.
        """
        test_row = start_row+2 #skip the first two settings used for gain ratio
        test_col = 2 #will read column 2 to identify the turning point of algorithm
        loop = True #loop test flag
        while loop ==True:
            #print(test_row,test_col)
            #record the cells below and above test cell
            cell1 = self.read_cell(test_row-1, test_col)
            cell2 = self.read_cell(test_row, test_col)
            cell3 = self.read_cell(test_row+1, test_col)
            #if they are all equal this must be the turninig point of the sequence
            if cell1==cell2 and cell1 ==cell3:
                loop = False
            else:
                test_row = test_row+1
        return (test_row,2*test_row-start_row) #center,last row
    
    def analysis(self):
        """
        Main analysis function. Reads averages and standard deviations to create GTC,ureal objects, by using the find_center function to determine starting and end points of a set.
        These are first sent to the split_set function to be seperated out for the different instruments, then each segment is sent to a ratio computation function which
        returns the gain ratio and lineariy ratio for that set.
        """
        end_point = float(self.read_cell(3,3))
        continue_analysis = True #flag for continuing to read through the data table
        start_row = int(float(self.read_cell(3,1)))-1 #wx starts at zero, initial starting row

        while continue_analysis ==True:
            center,last_row = self.find_center(start_row)
            print(start_row,center,last_row)
            if last_row>len(self.data): #last row is outside the table, last set is incomplete or not there at all
                print("last set not found")
                continue_analysis = False
            else:
                S_settings = [float(self.read_cell(i,self.S_setting_col)) for i in range(start_row,last_row+1)]
                X_settings = [float(self.read_cell(i,self.X_setting_col)) for i in range(start_row,last_row+1)]

                GTC_list = [GTC.ureal(float(self.read_cell(row,self.mean_col)),float(self.read_cell(row,self.std_col)),float(self.read_cell(row,self.DoF_col))-1) for row in range(start_row,last_row+1)]
                
                #get split up segments, then compute ratios individually and print then repeat.
                m_top,m_bottom,x_top,x_bottom,s0,s1 = self.split_set(GTC_list,S_settings)
                
                x_ratios = self.x_ratio(x_top,s0,s1)+self.x_ratio(x_bottom[::-1],s0,s1)
                m_ratios = self.m_ratio(m_top,s0,s1)+self.m_ratio(m_bottom[::-1],s0,s1) #join arrays of the three ratios, linearity, gain, fit ??

                ratios = x_ratios+m_ratios
                
                print("m ratios"+str([x.x for x in m_ratios]))
                print("x ratios"+str([x.x for x in x_ratios]))

                #print the ratios to the sheet
                self.PrintCol([x.label for x in ratios],start_row+1,self.mean_col+5,self.sh_results)
                self.PrintCol(["Ratio"]+[x.x for x in ratios],start_row,self.mean_col+6,self.sh_results)
                self.PrintCol(["STDEV"]+[x.u for x in ratios],start_row,self.mean_col+7,self.sh_results)
                self.PrintCol(["Effct. DoF"]+[x.df for x in ratios],start_row,self.mean_col+8,self.sh_results)
                
                #check if there is a next section:
                if last_row>=float(self.read_cell(3,3))-1:
                    continue_analysis = False

                else:
                    start_row = last_row+1

    

if __name__ =="__main__":
    a = Analyser('raw.simulated.xlsx','Sheet')
    a.analysis()
    a.Save('compare_murray.xlsx')
